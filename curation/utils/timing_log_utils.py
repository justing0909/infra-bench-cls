"""
timing_log_utils.py
-------------------
updates data/Infra-FM-timing-log.xlsx with per-region pipeline statistics:
file sizes, stage durations, acceptance rates, asset counts, and per-modality
tile counts.

the workbook is a running record filled in as regions complete, so it is
incomplete for regions that were curated before a given column existed.
COLUMN_MAP below is the contract between field names and sheet headers;
columns it names that the sheet lacks are appended automatically.
"""

from __future__ import annotations

import os
from typing import Any, Optional

import openpyxl


SHEET_NAME    = "Sheet1"
HEADER_ROW    = 2
DATA_START_ROW = 3

COLUMN_MAP = {
    # pre-processing
    "region":                    "sentinel data",
    "starting_file_size_kb":     "starting file size (KB)",
    "pre_filter_time_s":         "pre-filter time (s)",
    "power_only_file_size_mb":   'pre-filter ending file size (MB) [the "power_only" dataset]',
    "compression_time_s":        'compression to "non_solar_gen" time (s)',
    "collapsed_file_size_kb":    'non_solar_gen parquet size (KB)',
    "scanning_time_s":           'scanning time of non_solar_gen file (s)',

    # pipeline yield
    "gee_accept_pct":            "GEE Percent of accepted tiles (%)",
    "stac_accept_pct":           "STAC Percent of accepted tiles (%)",
    "qc_accept_pct":             "QC Percent of accepted tiles (%)",
    "triage_accept_pct":         "triage percent of accepted tiles (%)",

    # asset counts
    "assets_extracted":          "assets extracted",
    "assets_after_dedup":        "assets after dedup",

    # tile counts
    "total_tiles_fetched":       "total tiles fetched",
    "dataset_tiles":             "dataset tiles",

    # per-modality tile counts
    # these columns will be auto-created in the workbook if not present.
    "tiles_sentinel2_ms":        "tiles: sentinel2_ms",
    "tiles_sentinel1":           "tiles: sentinel1",
    "tiles_landsat_thermal":     "tiles: landsat_thermal",
    "tiles_naip":                "tiles: naip",
    "tiles_temporal":            "tiles: temporal stacks",

    # timing
    "total_time_elapsed_s":      "total time elapsed (s)",

    # run metadata
    "filter_preset":             "filter preset",
    "modalities":                "modalities",
    "temporal_stack":            "temporal stack",
}


def _normalize_region(region: str) -> str:
    return region.strip().lower().replace("_", " ")


def _find_header_columns(ws) -> dict[str, int]:
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=HEADER_ROW, column=col).value
        if isinstance(value, str):
            headers[value.strip()] = col
    return headers


def _find_or_create_region_row(ws, region: str) -> int:
    norm = _normalize_region(region)
    for row in range(DATA_START_ROW, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if isinstance(value, str) and _normalize_region(value) == norm:
            return row
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value = region
    return row


def _find_or_create_column(ws, header_name: str,
                            header_cols: dict) -> int:
    """
    returns the column index for header_name, creating it if it doesn't exist.
    used for per-modality columns that may not be in older workbooks.
    """
    if header_name in header_cols:
        return header_cols[header_name]

    new_col = ws.max_column + 1
    ws.cell(row=HEADER_ROW, column=new_col).value = header_name
    header_cols[header_name] = new_col
    return new_col


def update_timing_log(
    workbook_path : str,
    region        : str,
    **fields      : Any,
) -> None:
    """
    update one region row in the timing workbook with partial stats.
    fields not supplied are left untouched.

    supported field keys: see COLUMN_MAP above.
    per-modality fields (tiles_sentinel2_ms, tiles_sentinel1, etc.)
    auto-create columns in the workbook if they don't exist.
    """
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[SHEET_NAME]

    header_cols = _find_header_columns(ws)
    row = _find_or_create_region_row(ws, region)
    ws.cell(row=row, column=1).value = region

    for key, value in fields.items():
        if key not in COLUMN_MAP:
            raise KeyError(
                f"Unknown timing-log field: '{key}'. "
                f"Add it to COLUMN_MAP in timing_log_utils.py."
            )
        column_name = COLUMN_MAP[key]

        # per-modality and new metadata columns may not exist in older workbooks
        # — create them automatically rather than raising
        col = _find_or_create_column(ws, column_name, header_cols)
        ws.cell(row=row, column=col).value = value

    wb.save(workbook_path)


def update_modality_counts(
    workbook_path   : str,
    region          : str,
    modality_counts : dict,
    n_temporal      : int = 0,
) -> None:
    """
    convenience wrapper to write per-modality tile counts to the timing log.

    Parameters
    ----------
    modality_counts : dict mapping modality key (or combination string like
                      "sentinel2_ms+sentinel1") to tile count
    n_temporal      : number of tiles that have a temporal stack
    """
    fields = {}
    for modality_key, count in modality_counts.items():
        # map combination strings to individual modality columns
        for mod in ["sentinel2_ms", "sentinel1", "landsat_thermal", "naip"]:
            if mod in modality_key:
                field_key = f"tiles_{mod}"
                fields[field_key] = fields.get(field_key, 0) + count

    if n_temporal:
        fields["tiles_temporal"] = n_temporal

    if fields:
        update_timing_log(workbook_path, region, **fields)


def file_size_kb(path: str) -> Optional[float]:
    if not path or not os.path.exists(path):
        return None
    return round(os.path.getsize(path) / 1024.0, 2)


def file_size_mb(path: str) -> Optional[float]:
    if not path or not os.path.exists(path):
        return None
    return round(os.path.getsize(path) / (1024.0 * 1024.0), 2)